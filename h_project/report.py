import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
import win32com
from preprocessing import *
from predictions import *
import time

class make_report:
    def __init__(self, feed_data, conservative=0.1, moderate=0.07, liberal=0.04):
        self.feed_data = feed_data
        self.conservative = conservative
        self.moderate = moderate
        self.liberal = liberal

    def load_files(self, model_type='overall'):
        if model_type == 'overall':
            with open(folder + 'Best models/best_model_cc_threshes.txt', 'rb') as file:
                threshes = pickle.load(file)
            with open(folder + 'Best models/best_model_cc.pkl', 'rb') as file:
                model = pickle.load(file)

            self.threshes = threshes
            self.model = model
        elif model_type == 'mass_plus':
            with open(folder + 'Best models/best_model_mass_cc_threshes.txt', 'rb') as file:
                threshes_mass = pickle.load(file)
            with open(folder + 'Best models/best_model_mass_cc.pkl', 'rb') as file:
                model_mass = pickle.load(file)

            with open(folder + 'Best models/best_model_plus_cc_threshes.txt', 'rb') as file:
                threshes_plus = pickle.load(file)
            with open(folder + 'Best models/best_model_plus_cc.pkl', 'rb') as file:
                model_plus = pickle.load(file)

            self.threshes_mass = threshes_mass
            self.model_mass = model_mass
            self.threshes_plus = threshes_plus
            self.model_plus = model_plus

    def get_predictions(self, model_type='overall'):
        if model_type == 'overall':
            thresh_list = [self.conservative, self.moderate, self.liberal]
            predictions = pd.DataFrame([])
            predictions['Probability'] = pd.Series(self.model.predict_proba(self.feed_data)[:, 1]).apply \
                (lambda x: round(x, 2))
            for thresh, strategy in zip(thresh_list, ['Conservative', 'Moderate', 'Liberal']):
                predictions[strategy] = get_predictions(self.model, self.feed_data, thresh, self.threshes)
            predictions = predictions.loc[:, ['Conservative', 'Moderate', 'Liberal', 'Probability']]
            predictions.loc[:, ['Conservative', 'Moderate', 'Liberal']] = predictions.loc[:
            , ['Conservative', 'Moderate'
                  , 'Liberal']].replace \
                ({0: 'Non-default', 1: 'Default'})
            return predictions
        if model_type == 'mass_plus':
            thresh_list = [self.conservative, self.moderate, self.liberal]
            model_type = ['_mass', '_plus']
            mass_indices = self.feed_data[self.feed_data.isin([0])].index.values.tolist()
            plus_indices = self.feed_data[self.feed_data.isin([1, 2])].index.values.tolist()

            predictions = pd.DataFrame(np.zeros(self.feed_data.shape[0] * 4).reshape(-1, 4), index=self.feed_data.index,
                                       columns=['Probability', 'Conservative', 'Moderate', 'Liberal'])
            for thresh, strategy in zip(thresh_list, ['Probability', 'Conservative', 'Moderate', 'Liberal']):
                if strategy == 'Probability':
                    predictions.loc[mass_indices, strategy] = self.model_mass.predict_proba(
                        self.feed_data.loc[mass_indices])[:, 1]
                    predictions.loc[plus_indices, strategy] = self.model_plus.predict_proba(
                        self.feed_data.loc[plus_indices])[:, 1]
                else:
                    predictions.loc[mass_indices, strategy] = get_predictions(self.model_mass
                                                                              , self.feed_data.loc[mass_indices], thresh
                                                                              , self.threshes_mass)
                    predictions.loc[plus_indices, strategy] = get_predictions(self.model_plus
                                                                              , self.feed_data.loc[plus_indices], thresh
                                                                              , self.threshes_plus)

            predictions = predictions.loc[:, ['Conservative', 'Moderate', 'Liberal', 'Probability']]
            predictions.loc[:, ['Conservative', 'Moderate', 'Liberal']] = predictions.loc[:
            , ['Conservative', 'Moderate'
                  , 'Liberal']].replace \
                ({0: 'Non-default', 1: 'Default'})
            return predictions

    @property
    def overall_(self):
        self.load_files(model_type='overall')
        predictions = self.get_predictions(model_type='overall')
        return predictions

    @property
    def segmented_(self):
        self.load_files(model_type='mass_plus')
        predictions = self.get_predictions(model_type='mass_plus')
        return predictions


def find_column(sheet, value):
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row, col).value == value:
                destination_column = col
                break
    try:
        return destination_column
    except:
        raise ValueError('Value is not found on the sheet')


def find_row(sheet, value):
    for row in range(0, sheet.max_row + 1):
        for col in range(0, sheet.max_column + 1):
            if sheet.cell(row, col).value == value:
                destination_row = row
                break
    try:
        return destination_row
    except:
        raise ValueError('Value is not found on the sheet')


class update_preds:
    def __init__(self, new_file_path, sheet_name='Facility Data', save_path=None, password=None):
        self.new_file_path = new_file_path
        self.sheet_name = sheet_name
        self.save_path = save_path
        self.password = password

    def remove_password_xlsx(filename, password):
        xcl = win32com.client.Dispatch("Excel.Application")
        wb = xcl.Workbooks.Open(filename, False, False, None, password)
        xcl.DisplayAlerts = False
        wb.SaveAs(filename, None, '', '')
        xcl.Quit()

    def get_predictions(self):
        new_data = read_locked_xlsx(self.new_file_path, password=self.password)

        if self.password is not None:
            update_preds.remove_password_xlsx(self.new_file_path, self.password)

        xfile = openpyxl.load_workbook(self.new_file_path)
        sheet = xfile.get_sheet_by_name(self.sheet_name)
        alphabet_dict = dict(zip(range(1, 27), string.ascii_uppercase))

        col_num = find_column(sheet, 'Actual picture as a result of in-branch/RCS approvals')
        max_row = 0
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(i, col_num).value is not None:
                max_row += 1

        new_loans_count = sheet.max_row - max_row
        new_data = new_data.tail(new_loans_count)

        feed_data = get_feed_data(new_data, database='hsbc')

        rep = make_report(feed_data)

        self.overall_preds = rep.overall_
        self.segmented_preds = rep.segmented_
        self.feed_data = feed_data
        self.xfile = xfile
        self.sheet = sheet

    def add_preds(self, preds_type='Overall'):
        if preds_type == 'Overall':
            preds = self.overall_preds
        elif preds_type == 'Segmented':
            preds = self.segmented_preds

        col_num = find_column(self.sheet, 'Predictions provided by Fast Foundation (' + preds_type + ')')
        max_row = 0
        for i in range(1, self.sheet.max_row + 1):
            if self.sheet.cell(i, col_num).value is not None:
                max_row += 1
        for i, index in zip(range(max_row + 1, max_row + preds.shape[0] + 1), preds.index.values):
            for j, column in zip(range(col_num, col_num + preds.shape[1] + 1), preds.columns.values):
                self.sheet.cell(i, j).value = preds.loc[index, column]

        self.col_num = col_num
        self.max_row = max_row

        update_preds.format(self)

    def write_to_excel(self):

        if not self.save_path:
            self.xfile.save(self.new_file_path)
        else:
            self.xfile.save(self.save_path)


    def format(self):
        col_conservative = self.col_num
        col_moderate = self.col_num + 1
        col_liberal = self.col_num + 2
        col_probability = self.col_num + 3

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for row in range(3, self.sheet.max_row + 1):
            self.sheet.cell(row, col_probability).number_format = '0%'

            self.sheet.cell(row, col_conservative).border = thin_border
            self.sheet.cell(row, col_moderate).border = thin_border
            self.sheet.cell(row, col_liberal).border = thin_border
            self.sheet.cell(row, col_probability).border = thin_border

            self.sheet.cell(row, col_conservative).alignment = Alignment(horizontal='center')
            self.sheet.cell(row, col_moderate).alignment = Alignment(horizontal='center')
            self.sheet.cell(row, col_liberal).alignment = Alignment(horizontal='center')
            self.sheet.cell(row, col_probability).alignment = Alignment(horizontal='center')

            self.sheet.row_dimensions[row].height = 20

    def update(self):
        update_preds.get_predictions(self)
        update_preds.add_preds(self, preds_type='Overall')
        update_preds.add_preds(self, preds_type='Segmented')
        update_preds.write_to_excel(self)
